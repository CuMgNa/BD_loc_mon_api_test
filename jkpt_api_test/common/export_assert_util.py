# common/export_assert_util.py
"""二进制导出（xlsx）响应解析与结构断言。"""
from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from common.logger_util import key, sep
from common.yaml_util import read_expected_msg


@dataclass
class XlsxSheetSnapshot:
    sheet_name: str
    headers: list[str]
    data_row_count: int
    first_data_row: tuple[Any, ...] | None
    header_row: int = 1
    addr_column_values: list[str] | None = None


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(_normalize_cell(v) == "" for v in row)


def _trim_trailing_empty(values: list[str]) -> list[str]:
    result = list(values)
    while result and result[-1] == "":
        result.pop()
    return result


def _read_row(ws: Worksheet, row_idx: int) -> list[str]:
    """按 max_column 读取整行（避免 read_only 稀疏行只返回首格）。"""
    max_col = ws.max_column or 1
    return [_normalize_cell(ws.cell(row_idx, col).value) for col in range(1, max_col + 1)]


def _apply_merged_row_values(ws: Worksheet, row_idx: int, values: list[str]) -> list[str]:
    """横向合并单元格：将合并区左上角的值铺到整段合并列。"""
    row_vals = list(values)
    if not getattr(ws, "merged_cells", None):
        return _trim_trailing_empty(row_vals)

    for merged in ws.merged_cells.ranges:
        if merged.min_row != row_idx or merged.max_row != row_idx:
            continue
        if merged.min_col == merged.max_col:
            continue
        top_left = _normalize_cell(ws.cell(merged.min_row, merged.min_col).value)
        for col in range(merged.min_col, merged.max_col + 1):
            idx = col - 1
            if idx < len(row_vals) and not row_vals[idx]:
                row_vals[idx] = top_left
    return _trim_trailing_empty(row_vals)


def _find_header_row(
    ws: Worksheet,
    *,
    marker: str | None = None,
    expected_headers: list[str] | None = None,
    max_scan: int = 15,
) -> int:
    """在前 max_scan 行中定位表头行：优先完全匹配 expected，其次含 marker，否则非空列最多。"""
    if expected_headers:
        for row_idx in range(1, max_scan + 1):
            row = _apply_merged_row_values(ws, row_idx, _read_row(ws, row_idx))
            if row == expected_headers:
                return row_idx

    if marker:
        for row_idx in range(1, max_scan + 1):
            row = _apply_merged_row_values(ws, row_idx, _read_row(ws, row_idx))
            if marker in row:
                return row_idx

    best_row = 1
    best_count = 0
    for row_idx in range(1, max_scan + 1):
        row = _apply_merged_row_values(ws, row_idx, _read_row(ws, row_idx))
        count = sum(1 for cell in row if cell)
        if count > best_count:
            best_count = count
            best_row = row_idx
    return best_row


def parse_xlsx(
    content: bytes,
    *,
    header_marker: str | None = None,
    expected_headers: list[str] | None = None,
) -> XlsxSheetSnapshot:
    """解析 xlsx：自动定位表头行，统计数据行。"""
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    header_row = _find_header_row(
        ws,
        marker=header_marker,
        expected_headers=expected_headers,
    )
    headers = _apply_merged_row_values(ws, header_row, _read_row(ws, header_row))

    data_rows: list[tuple[Any, ...]] = []
    max_row = ws.max_row or header_row
    for row_idx in range(header_row + 1, max_row + 1):
        row = tuple(_read_row(ws, row_idx))
        if not _is_empty_row(row):
            data_rows.append(row)

    first_data_row = data_rows[0] if data_rows else None
    return XlsxSheetSnapshot(
        sheet_name=ws.title,
        headers=headers,
        data_row_count=len(data_rows),
        first_data_row=first_data_row,
        header_row=header_row,
    )


def _column_values(ws: Worksheet, headers: list[str], header_row: int, column_name: str) -> list[str]:
    try:
        col_idx = headers.index(column_name)
    except ValueError as exc:
        raise AssertionError(f"表头中未找到列 {column_name!r}，实际表头={headers}") from exc

    values: list[str] = []
    max_row = ws.max_row or header_row
    for row_idx in range(header_row + 1, max_row + 1):
        row = _read_row(ws, row_idx)
        if _is_empty_row(tuple(row)):
            continue
        cell = row[col_idx] if col_idx < len(row) else ""
        text = _normalize_cell(cell)
        if text:
            values.append(text)
    return values


def _parse_filename(content_disposition: str | None) -> str | None:
    if not content_disposition:
        return None
    match = re.search(r"filename\*?=(?:UTF-8''|\"?)([^\";]+)", content_disposition, re.IGNORECASE)
    if not match:
        return None
    return match.group(1).strip().strip('"')


def assert_xlsx_export_structure(
    *,
    case_name: str,
    content: bytes,
    expected: dict,
    addr_count: int | None = None,
    content_disposition: str | None = None,
) -> XlsxSheetSnapshot:
    """方案 B：校验 xlsx 魔数、表头、数据行数及 addr 列非空行数。"""
    min_size = expected.get("min_content_size", 512)
    assert len(content) >= min_size, (
        f"[{case_name}] 导出正文过小: 预期>={min_size}, 实际={len(content)}"
    )
    assert content[:2] == b"PK", (
        f"[{case_name}] 非 xlsx/zip 魔数: 前缀={content[:4]!r}"
    )

    expected_filename = expected.get("filename")
    if expected_filename:
        actual_filename = _parse_filename(content_disposition)
        assert actual_filename == expected_filename, (
            f"[{case_name}] 文件名不匹配: 预期={expected_filename}, 实际={actual_filename}"
        )

    expected_headers = expected.get("headers")
    addr_column = expected.get("addr_column")

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    header_row = _find_header_row(
        ws,
        marker=addr_column,
        expected_headers=expected_headers,
    )
    headers = _apply_merged_row_values(ws, header_row, _read_row(ws, header_row))

    data_rows: list[tuple[Any, ...]] = []
    max_row = ws.max_row or header_row
    for row_idx in range(header_row + 1, max_row + 1):
        row = tuple(_read_row(ws, row_idx))
        if not _is_empty_row(row):
            data_rows.append(row)

    snapshot = XlsxSheetSnapshot(
        sheet_name=ws.title,
        headers=headers,
        data_row_count=len(data_rows),
        first_data_row=data_rows[0] if data_rows else None,
        header_row=header_row,
    )

    if expected_headers is not None:
        assert snapshot.headers == expected_headers, (
            f"[{case_name}] 表头不匹配:\n"
            f"  表头行={header_row}\n"
            f"  预期={expected_headers}\n"
            f"  实际={snapshot.headers}"
        )

    min_rows = expected.get("min_rows")
    if min_rows is None and addr_count is not None:
        min_rows = addr_count
    if min_rows is not None:
        assert snapshot.data_row_count >= min_rows, (
            f"[{case_name}] 数据行不足: 预期>={min_rows}, 实际={snapshot.data_row_count}"
        )

    if addr_column and min_rows is not None:
        addr_values = _column_values(ws, snapshot.headers, header_row, addr_column)
        snapshot.addr_column_values = addr_values
        assert len(addr_values) >= min_rows, (
            f"[{case_name}] {addr_column!r} 列非空行不足: "
            f"预期>={min_rows}, 实际={len(addr_values)}"
        )

    sep(" 断言结果(xlsx 结构) ")
    key("Sheet", snapshot.sheet_name)
    key("表头行号", snapshot.header_row)
    key("表头", snapshot.headers)
    key("数据行数", snapshot.data_row_count)
    if snapshot.first_data_row is not None:
        key("首行数据", snapshot.first_data_row)
    if snapshot.addr_column_values is not None:
        key(f"{addr_column} 非空行数", len(snapshot.addr_column_values))

    return snapshot


def assert_export_response(
    *,
    case_name: str,
    response: Any,
    expected: dict,
    require_binary: bool = False,
    addr_count: int | None = None,
) -> None:
    """统一导出响应断言：
    - body 前缀像 JSON -> 断言业务 code/msg（若 require_binary=True 则直接失败）
    - 否则按二进制文件断言 HTTP/正文/xlsx 结构
    """
    raw = response.content or b""
    trimmed = raw.lstrip()
    looks_like_json = trimmed[:1] in (b"{", b"[")
    content_type = response.headers.get("Content-Type")
    content_disposition = response.headers.get("Content-Disposition")

    expected_http = expected.get("http_status")
    if expected_http is None:
        expected_http = expected.get("code")

    if looks_like_json:
        if expected_http is not None:
            assert response.status_code == expected_http, (
                f"[{case_name}] HTTP 状态码不匹配: 预期={expected_http}, 实际={response.status_code}"
            )
        try:
            payload = response.json()
        except Exception as exc:
            raise AssertionError(
                f"[{case_name}] 响应体看似 JSON，但解析失败: HTTP={response.status_code}, "
                f"Content-Type={content_type}, 前缀={trimmed[:120]!r}"
            ) from exc

        actual_code = payload.get("code")
        actual_msg = payload.get("msg", "")
        if require_binary:
            raise AssertionError(
                f"[{case_name}] 预期二进制导出文件，实际返回 JSON: "
                f"HTTP={response.status_code}, code={actual_code}, msg={actual_msg}, "
                f"Content-Type={content_type}"
            )

        expected_code = expected.get("code")
        if expected_code is not None:
            assert actual_code == expected_code, (
                f"[{case_name}] 业务码不匹配: 预期={expected_code}, 实际={actual_code}"
            )
        if "msg" in expected or "error_msg" in expected:
            expected_msg = read_expected_msg(expected)
            assert actual_msg == expected_msg, (
                f"[{case_name}] 业务消息不匹配: 预期={expected_msg}, 实际={actual_msg}"
            )
        return

    # 二进制分支
    assert expected_http is not None, (
        f"[{case_name}] 二进制响应需在 expected 中配置 http_status（或兼容字段 code）"
    )

    sep(" 断言结果(二进制导出) ")
    key("预期 HTTP 状态码", expected_http)
    key("实际 HTTP 状态码", response.status_code)
    key("Content-Type", content_type)
    key("Content-Disposition", content_disposition)
    key("响应体字节数", len(raw))

    assert response.status_code == expected_http, (
        f"[{case_name}] HTTP 状态码不匹配: 预期={expected_http}, 实际={response.status_code}"
    )
    assert len(raw) > 0, f"[{case_name}] 导出正文为空"

    if require_binary:
        # 防止“HTTP=200 但并未返回文件”
        assert raw[:2] == b"PK", (
            f"[{case_name}] 预期 xlsx 文件，实际非 zip/xlsx 魔数: 前缀={raw[:4]!r}"
        )
        if content_disposition:
            assert "attachment" in content_disposition.lower(), (
                f"[{case_name}] 预期文件下载头 attachment，实际={content_disposition}"
            )

    # 若配置了结构化预期，做 xlsx 深度断言
    if expected.get("headers") or expected.get("filename") or expected.get("addr_column"):
        assert_xlsx_export_structure(
            case_name=case_name,
            content=raw,
            expected=expected,
            addr_count=addr_count,
            content_disposition=content_disposition,
        )
